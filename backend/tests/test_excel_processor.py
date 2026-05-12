from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.app.excel_processor import (
    SummaryConfig,
    export_summary_workbook,
    preview_workbook,
    summarize_workbook,
)


def build_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws["A1"] = "销售明细表"
    headers = ["规格型号", "单位", "商品数量", "单价", "金额", "备注"]
    ws.append(headers)

    thin = Side(style="thin", color="D1D5DB")
    header_fill = PatternFill("solid", fgColor="217346")
    header_font = Font(color="FFFFFF", bold=True, name="Aptos")
    data_fill = PatternFill("solid", fgColor="E2F0D9")

    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 22
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ["US880-1R5-C02(1500w)", "台", 3, 693, 2079, "首批"],
        ["US880-1R5-C02(1500w)", "台", 1, 693, 693, "补货"],
        ["US880-4R0-C04(4000w)", "台", "bad", 837, 8370, "异常数量"],
        ["US880-4R0-C04(4000w)", "台", 2, 837, 1674, "补货"],
    ]

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=6):
        for cell in row:
            cell.fill = data_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
            if cell.column in (3, 4, 5):
                cell.number_format = "#,##0.00"

    summary = wb.create_sheet("其他")
    summary["A1"] = "另一个 sheet"
    wb.save(path)


def test_preview_workbook_returns_sheets_headers_rows_and_styles(tmp_path: Path) -> None:
    source = tmp_path / "销售明细.xlsx"
    build_sample_workbook(source)

    preview = preview_workbook(source, sheet_name="明细", header_row=2)

    assert preview.file_name == "销售明细.xlsx"
    assert preview.sheets == ["明细", "其他"]
    assert preview.headers[:5] == ["规格型号", "单位", "商品数量", "单价", "金额"]
    assert preview.rows[0]["规格型号"] == "US880-1R5-C02(1500w)"
    assert preview.columns[0].width_px > 150
    assert preview.header_style.background == "#217346"
    assert preview.data_style.background == "#E2F0D9"


def test_summarize_workbook_groups_by_selected_columns_and_sums_selected_columns(
    tmp_path: Path,
) -> None:
    source = tmp_path / "销售明细.xlsx"
    build_sample_workbook(source)
    config = SummaryConfig(
        sheet_name="明细",
        header_row=2,
        group_columns=["规格型号", "单位", "单价"],
        sum_columns=["商品数量", "金额"],
    )

    result = summarize_workbook(source, config)

    first = result.rows[0]
    assert first["规格型号"] == "US880-1R5-C02(1500w)"
    assert first["单位"] == "台"
    assert first["单价"] == 693
    assert first["商品数量"] == 4
    assert first["金额"] == 2772

    second = result.rows[1]
    assert second["规格型号"] == "US880-4R0-C04(4000w)"
    assert second["商品数量"] == 2
    assert second["金额"] == 10044
    assert "第 5 行的 商品数量 不是数字，已按 0 处理" in result.warnings


def test_summarize_workbook_rejects_missing_columns(tmp_path: Path) -> None:
    source = tmp_path / "销售明细.xlsx"
    build_sample_workbook(source)
    config = SummaryConfig(
        sheet_name="明细",
        header_row=2,
        group_columns=["不存在的列"],
        sum_columns=["金额"],
    )

    with pytest.raises(ValueError, match="不存在的列"):
        summarize_workbook(source, config)


def test_export_summary_workbook_uses_original_name_suffix_and_single_result_sheet(
    tmp_path: Path,
) -> None:
    source = tmp_path / "销售明细.xlsx"
    build_sample_workbook(source)
    config = SummaryConfig(
        sheet_name="明细",
        header_row=2,
        group_columns=["规格型号", "单位", "单价"],
        sum_columns=["商品数量", "金额"],
    )
    result = summarize_workbook(source, config)

    output = export_summary_workbook(
        result=result,
        original_file_name="销售明细.xlsx",
        output_dir=tmp_path,
    )

    assert output.name == "销售明细_汇总.xlsx"
    exported = load_workbook(output)
    assert exported.sheetnames == ["汇总结果"]
    ws = exported["汇总结果"]
    assert ws["A1"].value == "规格型号"
    assert ws["A1"].fill.fgColor.rgb == "00217346"
    assert ws["A2"].value == "US880-1R5-C02(1500w)"
    assert ws["D2"].value == 4
    assert ws.column_dimensions["A"].width == 28
    assert ws.row_dimensions[1].height == 26
