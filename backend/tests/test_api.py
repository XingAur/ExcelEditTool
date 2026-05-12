from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.main import app
from backend.tests.test_excel_processor import build_sample_workbook


client = TestClient(app)


def test_upload_preview_and_export_summary_file_name(tmp_path: Path) -> None:
    source = tmp_path / "销售明细.xlsx"
    build_sample_workbook(source)

    with source.open("rb") as handle:
        upload = client.post(
            "/api/workbooks",
            files={
                "file": (
                    "销售明细.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert upload.status_code == 200
    payload = upload.json()
    assert payload["fileName"] == "销售明细.xlsx"
    assert payload["sheets"] == ["明细", "其他"]
    assert payload["preview"]["headers"][0] == "规格型号"

    export = client.post(
        f"/api/workbooks/{payload['fileId']}/export",
        json={
            "sheet_name": "明细",
            "header_row": 2,
            "group_columns": ["规格型号", "单位", "单价"],
            "sum_columns": ["商品数量", "金额"],
        },
    )

    assert export.status_code == 200
    assert "%E9%94%80%E5%94%AE%E6%98%8E%E7%BB%86_%E6%B1%87%E6%80%BB.xlsx" in export.headers["content-disposition"]
    assert export.content[:2] == b"PK"


def test_export_edited_summary_results_keeps_multiple_result_sheets(tmp_path: Path) -> None:
    source = tmp_path / "销售明细.xlsx"
    build_sample_workbook(source)

    with source.open("rb") as handle:
        upload = client.post(
            "/api/workbooks",
            files={
                "file": (
                    "销售明细.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    file_id = upload.json()["fileId"]
    payload = {
        "results": [
            {
                "sheet_name": "上海",
                "header_row": 1,
                "columns": [
                    {"key": "型号", "index": 1, "letter": "A", "width": 12, "width_px": 89, "style": {}},
                    {"key": "金额", "index": 2, "letter": "B", "width": 12, "width_px": 89, "style": {}},
                ],
                "rows": [{"型号": "A", "金额": 99.5}],
                "warnings": [],
                "row_heights": {"1": 24, "2": 22},
                "header_style": {},
                "data_style": {},
            },
            {
                "sheet_name": "安徽",
                "header_row": 1,
                "columns": [
                    {"key": "型号", "index": 1, "letter": "A", "width": 12, "width_px": 89, "style": {}},
                    {"key": "金额", "index": 2, "letter": "B", "width": 12, "width_px": 89, "style": {}},
                ],
                "rows": [{"型号": "B", "金额": 18}],
                "warnings": [],
                "row_heights": {"1": 24, "2": 22},
                "header_style": {},
                "data_style": {},
            },
        ]
    }

    response = client.post(f"/api/workbooks/{file_id}/export-results", json=payload)

    assert response.status_code == 200
    target = tmp_path / "edited.xlsx"
    target.write_bytes(response.content)
    wb = load_workbook(target)
    assert wb.sheetnames == ["上海", "安徽"]
    ws = wb["上海"]
    assert ws["B2"].value == 99.5
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:B2"
    assert ws.row_dimensions[1].height == 24
    assert ws.column_dimensions["A"].width == 12
    assert ws["A1"].fill.fill_type == "solid"
    assert ws["A1"].font.bold
    assert ws["A1"].border.left.style == "thin"


def test_guide_preference_is_persisted(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_EDIT_TOOL_CONFIG_DIR", str(tmp_path))

    initial = client.get("/api/preferences/guide")
    assert initial.status_code == 200
    assert initial.json() == {"guideHidden": False}

    saved = client.post("/api/preferences/guide", json={"guideHidden": True})
    assert saved.status_code == 200
    assert saved.json() == {"guideHidden": True}

    loaded = client.get("/api/preferences/guide")
    assert loaded.status_code == 200
    assert loaded.json() == {"guideHidden": True}
