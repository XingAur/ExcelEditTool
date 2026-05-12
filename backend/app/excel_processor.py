from __future__ import annotations

from collections import OrderedDict
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field


class CellStyle(BaseModel):
    background: str | None = None
    color: str | None = None
    bold: bool = False
    horizontal: str | None = None
    vertical: str | None = None
    number_format: str | None = None
    border_color: str | None = None
    border_style: str | None = None
    font_name: str | None = None
    font_size: float | None = None


class PreviewColumn(BaseModel):
    key: str
    index: int
    letter: str
    width: float
    width_px: int
    style: CellStyle


class SheetPreview(BaseModel):
    file_name: str
    sheets: list[str]
    sheet_name: str
    header_row: int
    headers: list[str]
    rows: list[dict[str, Any]]
    columns: list[PreviewColumn]
    row_heights: dict[int, float]
    header_style: CellStyle
    data_style: CellStyle


class SummaryConfig(BaseModel):
    sheet_name: str
    header_row: int = Field(ge=1)
    group_columns: list[str]
    sum_columns: list[str]


class SummaryResult(BaseModel):
    sheet_name: str
    header_row: int
    columns: list[PreviewColumn]
    rows: list[dict[str, Any]]
    warnings: list[str]
    row_heights: dict[int, float]
    header_style: CellStyle
    data_style: CellStyle


class SummaryExportRequest(BaseModel):
    results: list[SummaryResult]


def preview_workbook(
    workbook_path: str | Path,
    sheet_name: str | None = None,
    header_row: int = 1,
    max_rows: int = 200,
) -> SheetPreview:
    path = Path(workbook_path)
    wb = load_workbook(path, data_only=True)
    active_sheet_name = sheet_name or wb.sheetnames[0]
    if active_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet 不存在: {active_sheet_name}")

    ws = wb[active_sheet_name]
    headers = _read_headers(ws, header_row)
    columns = _read_columns(ws, headers, header_row)
    rows: list[dict[str, Any]] = []

    for row_number in range(header_row + 1, min(ws.max_row, header_row + max_rows) + 1):
        row: dict[str, Any] = {}
        has_value = False
        for column_index, header in enumerate(headers, start=1):
            value = _json_value(ws.cell(row=row_number, column=column_index).value)
            row[header] = value
            has_value = has_value or value not in (None, "")
        if has_value:
            rows.append(row)

    return SheetPreview(
        file_name=path.name,
        sheets=wb.sheetnames,
        sheet_name=active_sheet_name,
        header_row=header_row,
        headers=headers,
        rows=rows,
        columns=columns,
        row_heights=_read_row_heights(ws, header_row),
        header_style=_style_from_cell(ws.cell(row=header_row, column=1)),
        data_style=_style_from_cell(ws.cell(row=header_row + 1, column=1)),
    )


def detect_header_row(workbook_path: str | Path, sheet_name: str | None = None) -> int:
    wb = load_workbook(workbook_path, data_only=True)
    active_sheet_name = sheet_name or wb.sheetnames[0]
    if active_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet 不存在: {active_sheet_name}")

    ws = wb[active_sheet_name]
    best_row = 1
    best_count = 0
    for row_number in range(1, min(ws.max_row, 10) + 1):
        count = sum(
            1
            for cell in ws[row_number]
            if cell.value not in (None, "")
        )
        if count > best_count:
            best_row = row_number
            best_count = count
    return best_row


def summarize_workbook(workbook_path: str | Path, config: SummaryConfig) -> SummaryResult:
    wb = load_workbook(workbook_path, data_only=True)
    if config.sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet 不存在: {config.sheet_name}")

    ws = wb[config.sheet_name]
    headers = _read_headers(ws, config.header_row)
    _ensure_columns_exist(headers, config.group_columns + config.sum_columns)

    header_to_index = {header: index for index, header in enumerate(headers, start=1)}
    output_headers = [*config.group_columns, *config.sum_columns]
    output_columns = _read_columns(ws, output_headers, config.header_row, header_to_index)
    groups: OrderedDict[tuple[str, ...], dict[str, Any]] = OrderedDict()
    warnings: list[str] = []

    for row_number in range(config.header_row + 1, ws.max_row + 1):
        source_row = {
            header: _json_value(ws.cell(row=row_number, column=index).value)
            for header, index in header_to_index.items()
        }
        if not any(value not in (None, "") for value in source_row.values()):
            continue

        group_key = tuple(_key_value(source_row[column]) for column in config.group_columns)
        if group_key not in groups:
            groups[group_key] = {
                column: source_row[column] for column in config.group_columns
            }
            for column in config.sum_columns:
                groups[group_key][column] = Decimal("0")

        for column in config.sum_columns:
            value = _decimal_value(source_row[column])
            if value is None:
                warnings.append(f"第 {row_number} 行的 {column} 不是数字，已按 0 处理")
                value = Decimal("0")
            groups[group_key][column] += value

    rows = []
    for grouped_row in groups.values():
        rows.append(
            {
                column: _number_value(grouped_row[column])
                if column in config.sum_columns
                else grouped_row[column]
                for column in output_headers
            }
        )

    return SummaryResult(
        sheet_name="汇总结果",
        header_row=1,
        columns=output_columns,
        rows=rows,
        warnings=warnings,
        row_heights=_read_row_heights(ws, config.header_row),
        header_style=_style_from_cell(ws.cell(row=config.header_row, column=1)),
        data_style=_style_from_cell(ws.cell(row=config.header_row + 1, column=1)),
    )


def export_summary_workbook(
    result: SummaryResult,
    original_file_name: str,
    output_dir: str | Path,
) -> Path:
    return export_summary_results_workbook([result], original_file_name, output_dir)


def export_summary_results_workbook(
    results: list[SummaryResult],
    original_file_name: str,
    output_dir: str | Path,
) -> Path:
    if not results:
        raise ValueError("没有可导出的汇总结果")

    output_path = Path(output_dir) / _summary_file_name(original_file_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_titles: set[str] = set()
    for result in results:
        title = _unique_sheet_title(result.sheet_name or "汇总结果", used_titles)
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        _write_summary_sheet(ws, result)

    wb.save(output_path)
    return output_path


def _write_summary_sheet(ws: Any, result: SummaryResult) -> None:
    header_height = result.row_heights.get(1) or result.row_heights.get(result.header_row) or 24
    data_height = result.row_heights.get(2) or 22
    header_style = _style_with_defaults(result.header_style, header=True)
    data_style = _style_with_defaults(result.data_style, header=False)

    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = header_height

    for column_index, column in enumerate(result.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column.key)
        _apply_style(cell, header_style)
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = max(6, min(float(column.width or 12), 60))

    for row_offset, row in enumerate(result.rows, start=2):
        ws.row_dimensions[row_offset].height = data_height
        for column_index, column in enumerate(result.columns, start=1):
            cell = ws.cell(row=row_offset, column=column_index, value=row.get(column.key))
            _apply_style(cell, _style_with_defaults(_merge_styles(data_style, column.style), header=False))

    if result.columns:
        last_column = get_column_letter(len(result.columns))
        last_row = max(1, len(result.rows) + 1)
        ws.auto_filter.ref = f"A1:{last_column}{last_row}"


def _unique_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    safe = "".join(char for char in raw_title if char not in r'[]:*?/\\').strip()[:31]
    safe = safe or "汇总结果"
    if safe not in used_titles:
        return safe
    base = safe[:28]
    index = 2
    while f"{base}_{index}" in used_titles:
        index += 1
    return f"{base}_{index}"


def _summary_file_name(original_file_name: str) -> str:
    stem = Path(original_file_name).stem
    return f"{stem}_汇总.xlsx"


def _read_headers(ws: Any, header_row: int) -> list[str]:
    headers: list[str] = []
    for index in range(1, ws.max_column + 1):
        raw_value = ws.cell(row=header_row, column=index).value
        header = str(raw_value).strip() if raw_value not in (None, "") else f"列{index}"
        headers.append(header)
    return headers


def _read_columns(
    ws: Any,
    headers: list[str],
    header_row: int,
    header_to_index: dict[str, int] | None = None,
) -> list[PreviewColumn]:
    columns: list[PreviewColumn] = []
    header_to_index = header_to_index or {
        header: index for index, header in enumerate(headers, start=1)
    }

    for output_index, header in enumerate(headers, start=1):
        source_index = header_to_index[header]
        letter = get_column_letter(source_index)
        width = ws.column_dimensions[letter].width or 10
        columns.append(
            PreviewColumn(
                key=header,
                index=output_index,
                letter=get_column_letter(output_index),
                width=width,
                width_px=_excel_width_to_px(width),
                style=_style_from_cell(ws.cell(row=header_row + 1, column=source_index)),
            )
        )
    return columns


def _read_row_heights(ws: Any, header_row: int) -> dict[int, float]:
    return {
        1: ws.row_dimensions[header_row].height or 24,
        2: ws.row_dimensions[header_row + 1].height or 22,
    }


def _ensure_columns_exist(headers: list[str], requested: list[str]) -> None:
    missing = [column for column in requested if column not in headers]
    if missing:
        raise ValueError(f"列不存在: {', '.join(missing)}")


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return _number_value(value)
    return value


def _key_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decimal_value(value: Any) -> Decimal | None:
    if value in (None, ""):
        return Decimal("0")
    try:
        normalized = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None


def _number_value(value: Any) -> int | float | Any:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    return value


def _style_from_cell(cell: Cell) -> CellStyle:
    return CellStyle(
        background=_color_to_hex(cell.fill.fgColor.rgb if cell.fill else None),
        color=_color_to_hex(cell.font.color.rgb if cell.font and cell.font.color and cell.font.color.type == "rgb" else None),
        bold=bool(cell.font.bold) if cell.font else False,
        horizontal=cell.alignment.horizontal if cell.alignment else None,
        vertical=cell.alignment.vertical if cell.alignment else None,
        number_format=cell.number_format,
        border_color=_border_color(cell),
        border_style=_border_style(cell),
        font_name=cell.font.name if cell.font else None,
        font_size=float(cell.font.sz) if cell.font and cell.font.sz else None,
    )


def _color_to_hex(value: str | None) -> str | None:
    if not value or value == "00000000":
        return None
    return f"#{value[-6:]}"


def _border_color(cell: Cell) -> str | None:
    side = cell.border.left or cell.border.right or cell.border.top or cell.border.bottom
    if not side or not side.color:
        return None
    return _color_to_hex(side.color.rgb)


def _border_style(cell: Cell) -> str | None:
    for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom):
        if side and side.style:
            return side.style
    return None


def _merge_styles(base: CellStyle, override: CellStyle) -> CellStyle:
    return CellStyle(
        background=override.background or base.background,
        color=override.color or base.color,
        bold=override.bold or base.bold,
        horizontal=override.horizontal or base.horizontal,
        vertical=override.vertical or base.vertical,
        number_format=override.number_format or base.number_format,
        border_color=override.border_color or base.border_color,
        border_style=override.border_style or base.border_style,
        font_name=override.font_name or base.font_name,
        font_size=override.font_size or base.font_size,
    )


def _style_with_defaults(style: CellStyle, *, header: bool) -> CellStyle:
    return CellStyle(
        background=style.background or ("#F3F3F3" if header else "#FFFFFF"),
        color=style.color or "#242424",
        bold=style.bold or header,
        horizontal=style.horizontal or ("center" if header else None),
        vertical=style.vertical or "center",
        number_format=style.number_format,
        border_color=style.border_color or "#D9D9D9",
        border_style=style.border_style or "thin",
        font_name=style.font_name or "Aptos",
        font_size=style.font_size or 11,
    )


def _apply_style(cell: Cell, style: CellStyle) -> None:
    if style.background:
        cell.fill = PatternFill("solid", fgColor=style.background.replace("#", ""))
    if style.font_name or style.font_size or style.color or style.bold:
        cell.font = Font(
            name=style.font_name or "Aptos",
            size=style.font_size,
            color=style.color.replace("#", "") if style.color else None,
            bold=style.bold,
        )
    cell.alignment = Alignment(
        horizontal=style.horizontal,
        vertical=style.vertical or "center",
    )
    if style.number_format:
        cell.number_format = style.number_format
    if style.border_style or style.border_color:
        color = style.border_color.replace("#", "") if style.border_color else "D1D5DB"
        side = Side(style=style.border_style or "thin", color=color)
        cell.border = Border(left=copy(side), right=copy(side), top=copy(side), bottom=copy(side))


def _excel_width_to_px(width: float) -> int:
    return max(48, int(width * 7 + 5))
